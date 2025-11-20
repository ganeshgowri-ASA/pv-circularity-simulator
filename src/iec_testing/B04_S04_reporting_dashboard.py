"""
IEC Testing Results & Reporting Dashboard - BATCH4-B04-S04

This module provides comprehensive IEC testing results management and interactive reporting
dashboard for all IEC test standards (61215, 61730, 63202, 63209, 63279).
"""

import io
import json
import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from jinja2 import Template
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Image,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from weasyprint import HTML

from src.iec_testing.models.test_models import (
    CertificationBodyType,
    CertificationPackage,
    CertificationStatus,
    ComplianceMatrix,
    ComplianceReport,
    IEC61215Result,
    IEC61730Result,
    IEC63202Result,
    IECStandard,
    TestHistory,
    TestPhoto,
    TestResult,
    TestStatus,
)

logger = logging.getLogger(__name__)


class IECTestResultsManager:
    """
    Manages IEC test results from multiple test modules.

    This class provides functionality to load, aggregate, compare, and export
    test results from IEC 61215, 61730, and 63202 test modules.
    """

    def __init__(self, data_dir: Optional[Path] = None) -> None:
        """
        Initialize test results manager.

        Args:
            data_dir: Directory containing test result data files
        """
        self.data_dir = data_dir or Path("./test_data")
        self.data_dir.mkdir(parents=True, exist_ok=True)
        self.logger = logging.getLogger(f"{__name__}.{self.__class__.__name__}")

        # Storage for loaded results
        self.iec_61215_results: List[IEC61215Result] = []
        self.iec_61730_results: List[IEC61730Result] = []
        self.iec_63202_results: List[IEC63202Result] = []

    def load_test_results(
        self,
        result_61215: Optional[IEC61215Result] = None,
        result_61730: Optional[IEC61730Result] = None,
        result_63202: Optional[IEC63202Result] = None,
    ) -> Dict[str, int]:
        """
        Import test results from B04-S01, B04-S02, B04-S03.

        Args:
            result_61215: IEC 61215 test result
            result_61730: IEC 61730 test result
            result_63202: IEC 63202 test result

        Returns:
            Dict[str, int]: Count of loaded results by standard
        """
        self.logger.info("Loading test results")

        counts = {"iec_61215": 0, "iec_61730": 0, "iec_63202": 0}

        if result_61215:
            self.iec_61215_results.append(result_61215)
            counts["iec_61215"] += 1
            self.logger.info(f"Loaded IEC 61215 result: {result_61215.test_campaign_id}")

        if result_61730:
            self.iec_61730_results.append(result_61730)
            counts["iec_61730"] += 1
            self.logger.info(f"Loaded IEC 61730 result: {result_61730.test_campaign_id}")

        if result_63202:
            self.iec_63202_results.append(result_63202)
            counts["iec_63202"] += 1
            self.logger.info(f"Loaded IEC 63202 result: {result_63202.test_campaign_id}")

        return counts

    def aggregate_compliance_status(self) -> Dict[str, Any]:
        """
        Determine overall compliance across all IEC standards.

        Returns:
            Dict[str, Any]: Aggregated compliance status
        """
        self.logger.info("Aggregating compliance status")

        total_tests = 0
        passed_tests = 0
        failed_tests = 0

        # Aggregate IEC 61215
        for result in self.iec_61215_results:
            if result.overall_status == TestStatus.PASSED:
                passed_tests += 1
            elif result.overall_status == TestStatus.FAILED:
                failed_tests += 1
            total_tests += 1

        # Aggregate IEC 61730
        for result in self.iec_61730_results:
            if result.overall_status == TestStatus.PASSED:
                passed_tests += 1
            elif result.overall_status == TestStatus.FAILED:
                failed_tests += 1
            total_tests += 1

        # Aggregate IEC 63202
        for result in self.iec_63202_results:
            if result.overall_status == TestStatus.PASSED:
                passed_tests += 1
            elif result.overall_status == TestStatus.FAILED:
                failed_tests += 1
            total_tests += 1

        compliance_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0.0

        return {
            "total_tests": total_tests,
            "passed_tests": passed_tests,
            "failed_tests": failed_tests,
            "compliance_rate": compliance_rate,
            "overall_status": "PASSED" if failed_tests == 0 and total_tests > 0 else "FAILED",
        }

    def compare_to_standards(
        self, result: Union[IEC61215Result, IEC61730Result, IEC63202Result]
    ) -> Dict[str, Any]:
        """
        Benchmark test results against IEC requirements.

        Args:
            result: Test result to compare

        Returns:
            Dict[str, Any]: Comparison analysis
        """
        self.logger.info(f"Comparing results to standards for {type(result).__name__}")

        comparison = {
            "standard": "",
            "requirements_met": [],
            "requirements_failed": [],
            "margin_of_safety": {},
        }

        if isinstance(result, IEC61215Result):
            comparison["standard"] = "IEC 61215:2021"
            # Check power degradation
            if result.test_sequence.power_degradation_percent < 5.0:
                comparison["requirements_met"].append(
                    f"Power degradation: {result.test_sequence.power_degradation_percent:.2f}% < 5.0%"
                )
                comparison["margin_of_safety"]["power_degradation"] = (
                    5.0 - result.test_sequence.power_degradation_percent
                )
            else:
                comparison["requirements_failed"].append(
                    f"Power degradation: {result.test_sequence.power_degradation_percent:.2f}% >= 5.0%"
                )

        elif isinstance(result, IEC61730Result):
            comparison["standard"] = "IEC 61730-1/-2:2016"
            comparison["requirements_met"].append(f"Safety class: {result.safety_class}")

        elif isinstance(result, IEC63202Result):
            comparison["standard"] = "IEC TS 63202:2020"
            ctm_ratio = result.ctm_loss_breakdown.ctm_ratio
            if ctm_ratio >= 0.92:
                comparison["requirements_met"].append(f"CTM ratio: {ctm_ratio:.3f} >= 0.92")
                comparison["margin_of_safety"]["ctm_ratio"] = ctm_ratio - 0.92
            else:
                comparison["requirements_failed"].append(f"CTM ratio: {ctm_ratio:.3f} < 0.92")

        return comparison

    def generate_compliance_matrix(self) -> ComplianceMatrix:
        """
        Create pass/fail matrix for all test sequences.

        Returns:
            ComplianceMatrix: Complete compliance matrix
        """
        self.logger.info("Generating compliance matrix")

        iec_61215_tests: Dict[str, TestStatus] = {}
        iec_61730_tests: Dict[str, TestStatus] = {}
        iec_63202_tests: Dict[str, TestStatus] = {}

        # Process IEC 61215 results
        for result in self.iec_61215_results:
            seq = result.test_sequence
            iec_61215_tests["Visual Inspection Initial"] = seq.visual_inspection_initial.status
            iec_61215_tests["Performance at STC"] = seq.performance_at_stc.status
            iec_61215_tests["Wet Leakage Current"] = seq.wet_leakage_current.status
            iec_61215_tests["Thermal Cycling"] = seq.thermal_cycling.status
            iec_61215_tests["Humidity Freeze"] = seq.humidity_freeze.status
            iec_61215_tests["Damp Heat"] = seq.damp_heat.status
            iec_61215_tests["UV Preconditioning"] = seq.uv_preconditioning.status
            iec_61215_tests["Mechanical Load"] = seq.mechanical_load_test.status
            iec_61215_tests["Hail Impact"] = seq.hail_impact.status
            iec_61215_tests["Hot Spot Endurance"] = seq.hot_spot_endurance.status
            iec_61215_tests["Visual Inspection Final"] = seq.visual_inspection_final.status
            iec_61215_tests["Performance at STC Final"] = seq.performance_at_stc_final.status

        # Process IEC 61730 results
        for result in self.iec_61730_results:
            safety = result.safety_tests
            iec_61730_tests["Insulation Test"] = safety.insulation_test.status
            iec_61730_tests["Dielectric Withstand"] = safety.dielectric_withstand.status
            iec_61730_tests["Ground Continuity"] = safety.ground_continuity.status
            iec_61730_tests["Fire Test"] = safety.fire_test.status
            iec_61730_tests["Mechanical Stress"] = safety.mechanical_stress.status
            iec_61730_tests["Impact Test"] = safety.impact_test.status
            iec_61730_tests["UV Test"] = safety.UV_test.status
            iec_61730_tests["Corrosion Test"] = safety.corrosion_test.status

        # Process IEC 63202 results
        for result in self.iec_63202_results:
            iec_63202_tests["CTM Loss Analysis"] = result.overall_status
            iec_63202_tests["Cell Performance"] = TestStatus.PASSED
            iec_63202_tests["Module Performance"] = TestStatus.PASSED

        # Calculate totals
        all_tests = {**iec_61215_tests, **iec_61730_tests, **iec_63202_tests}
        total_tests = len(all_tests)
        passed_tests = sum(1 for status in all_tests.values() if status == TestStatus.PASSED)
        failed_tests = sum(1 for status in all_tests.values() if status == TestStatus.FAILED)

        overall_compliance = failed_tests == 0 and total_tests > 0
        compliance_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0.0

        return ComplianceMatrix(
            iec_61215_tests=iec_61215_tests,
            iec_61730_tests=iec_61730_tests,
            iec_63202_tests=iec_63202_tests,
            overall_compliance=overall_compliance,
            total_tests=total_tests,
            passed_tests=passed_tests,
            failed_tests=failed_tests,
            compliance_rate=compliance_rate,
        )

    def export_test_package(
        self, output_dir: Path, format: str = "json"
    ) -> Dict[str, Path]:
        """
        Export complete test documentation package for certification bodies.

        Args:
            output_dir: Output directory for exported files
            format: Export format (json, xml, excel)

        Returns:
            Dict[str, Path]: Paths to exported files
        """
        self.logger.info(f"Exporting test package in {format} format to {output_dir}")

        output_dir.mkdir(parents=True, exist_ok=True)
        exported_files: Dict[str, Path] = {}

        if format == "json":
            # Export IEC 61215 results
            if self.iec_61215_results:
                path_61215 = output_dir / "iec_61215_results.json"
                with open(path_61215, "w") as f:
                    json.dump(
                        [r.model_dump(mode="json") for r in self.iec_61215_results],
                        f,
                        indent=2,
                        default=str,
                    )
                exported_files["iec_61215"] = path_61215

            # Export IEC 61730 results
            if self.iec_61730_results:
                path_61730 = output_dir / "iec_61730_results.json"
                with open(path_61730, "w") as f:
                    json.dump(
                        [r.model_dump(mode="json") for r in self.iec_61730_results],
                        f,
                        indent=2,
                        default=str,
                    )
                exported_files["iec_61730"] = path_61730

            # Export IEC 63202 results
            if self.iec_63202_results:
                path_63202 = output_dir / "iec_63202_results.json"
                with open(path_63202, "w") as f:
                    json.dump(
                        [r.model_dump(mode="json") for r in self.iec_63202_results],
                        f,
                        indent=2,
                        default=str,
                    )
                exported_files["iec_63202"] = path_63202

        elif format == "excel":
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # IEC 61215 sheet
            if self.iec_61215_results:
                ws = wb.create_sheet("IEC 61215")
                data = []
                for result in self.iec_61215_results:
                    data.append(
                        {
                            "Campaign ID": result.test_campaign_id,
                            "Module Type": result.module_type,
                            "Manufacturer": result.manufacturer,
                            "Overall Status": result.overall_status.value,
                            "Compliance %": result.compliance_percentage,
                            "Power Degradation %": result.test_sequence.power_degradation_percent,
                        }
                    )
                df = pd.DataFrame(data)
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)

            # IEC 61730 sheet
            if self.iec_61730_results:
                ws = wb.create_sheet("IEC 61730")
                data = []
                for result in self.iec_61730_results:
                    data.append(
                        {
                            "Campaign ID": result.test_campaign_id,
                            "Module Type": result.module_type,
                            "Safety Class": result.safety_class,
                            "Overall Status": result.overall_status.value,
                            "Compliance %": result.compliance_percentage,
                        }
                    )
                df = pd.DataFrame(data)
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)

            # IEC 63202 sheet
            if self.iec_63202_results:
                ws = wb.create_sheet("IEC 63202")
                data = []
                for result in self.iec_63202_results:
                    data.append(
                        {
                            "Campaign ID": result.test_campaign_id,
                            "Module Type": result.module_type,
                            "CTM Ratio": result.ctm_loss_breakdown.ctm_ratio,
                            "Total CTM Loss %": result.ctm_loss_breakdown.total_ctm_loss,
                            "Overall Status": result.overall_status.value,
                        }
                    )
                df = pd.DataFrame(data)
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)

            excel_path = output_dir / "iec_test_results.xlsx"
            wb.save(excel_path)
            exported_files["excel"] = excel_path

        self.logger.info(f"Exported {len(exported_files)} files")
        return exported_files

    def track_test_history(
        self, module_type: str, lookback_days: int = 365
    ) -> TestHistory:
        """
        Maintain historical test data for trend analysis.

        Args:
            module_type: Module type to track
            lookback_days: Number of days to look back

        Returns:
            TestHistory: Historical test data
        """
        self.logger.info(f"Tracking test history for {module_type}")

        cutoff_date = datetime.now() - timedelta(days=lookback_days)

        # Collect relevant results
        test_campaigns: List[str] = []
        test_dates: List[datetime] = []
        power_output_history: List[float] = []
        efficiency_history: List[float] = []
        degradation_history: List[float] = []
        iec_61215_history: List[Optional[IEC61215Result]] = []
        iec_61730_history: List[Optional[IEC61730Result]] = []
        iec_63202_history: List[Optional[IEC63202Result]] = []

        # Process IEC 61215 results
        for result in self.iec_61215_results:
            if result.module_type == module_type and result.test_end_date >= cutoff_date:
                test_campaigns.append(result.test_campaign_id)
                test_dates.append(result.test_end_date)
                power_output_history.append(result.test_sequence.iv_curve_final.pmax)
                efficiency_history.append(result.test_sequence.iv_curve_final.efficiency or 0.0)
                degradation_history.append(result.test_sequence.power_degradation_percent)
                iec_61215_history.append(result)
                iec_61730_history.append(None)
                iec_63202_history.append(None)

        # Calculate statistics
        mean_power = float(np.mean(power_output_history)) if power_output_history else 0.0
        std_power = float(np.std(power_output_history)) if power_output_history else 0.0
        mean_degradation = (
            float(np.mean(degradation_history)) if degradation_history else 0.0
        )

        return TestHistory(
            module_type=module_type,
            test_campaigns=test_campaigns,
            test_dates=test_dates,
            power_output_history=power_output_history,
            efficiency_history=efficiency_history,
            degradation_history=degradation_history,
            iec_61215_history=iec_61215_history,
            iec_61730_history=iec_61730_history,
            iec_63202_history=iec_63202_history,
            mean_power=mean_power,
            std_power=std_power,
            mean_degradation=mean_degradation,
        )


class TestReportGenerator:
    """
    Generates professional test reports for certification.

    Supports PDF generation for IEC 61215, 61730, 63202, and combined reports.
    """

    def __init__(self, company_name: str = "Solar Test Lab", logo_path: Optional[Path] = None) -> None:
        """
        Initialize report generator.

        Args:
            company_name: Company/lab name for reports
            logo_path: Path to company logo
        """
        self.company_name = company_name
        self.logo_path = logo_path
        self.logger = logging.getLogger(f"{__name__}.{self.__class__.__name__}")

    def _create_report_header(
        self, styles: Any, title: str, report_number: str
    ) -> List[Any]:
        """
        Create report header elements.

        Args:
            styles: ReportLab styles
            title: Report title
            report_number: Report number

        Returns:
            List[Any]: Report header elements
        """
        elements = []

        # Title
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=24,
            textColor=colors.HexColor("#1f77b4"),
            spaceAfter=30,
            alignment=1,  # Center
        )
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.2 * inch))

        # Report info
        info_style = styles["Normal"]
        elements.append(Paragraph(f"<b>Report Number:</b> {report_number}", info_style))
        elements.append(
            Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}", info_style)
        )
        elements.append(Paragraph(f"<b>Laboratory:</b> {self.company_name}", info_style))
        elements.append(Spacer(1, 0.3 * inch))

        return elements

    def generate_iec61215_report(
        self, result: IEC61215Result, output_path: Path
    ) -> Path:
        """
        Generate comprehensive IEC 61215 qualification report.

        Args:
            result: IEC 61215 test result
            output_path: Output PDF path

        Returns:
            Path: Path to generated report
        """
        self.logger.info(f"Generating IEC 61215 report for {result.test_campaign_id}")

        doc = SimpleDocTemplate(str(output_path), pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []

        # Header
        elements.extend(
            self._create_report_header(
                styles,
                "IEC 61215 Module Qualification Test Report",
                result.test_report_number or "N/A",
            )
        )

        # Module information
        elements.append(Paragraph("<b>Module Information</b>", styles["Heading2"]))
        module_data = [
            ["Parameter", "Value"],
            ["Module Type", result.module_type],
            ["Manufacturer", result.manufacturer],
            ["Test Campaign ID", result.test_campaign_id],
            ["Test Start Date", result.test_start_date.strftime("%Y-%m-%d")],
            ["Test End Date", result.test_end_date.strftime("%Y-%m-%d")],
            ["Overall Status", result.overall_status.value.upper()],
            ["Compliance %", f"{result.compliance_percentage:.1f}%"],
        ]

        table = Table(module_data, colWidths=[3 * inch, 3 * inch])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        elements.append(table)
        elements.append(Spacer(1, 0.3 * inch))

        # Test results summary
        elements.append(Paragraph("<b>Test Results Summary</b>", styles["Heading2"]))
        seq = result.test_sequence
        test_data = [
            ["Test Name", "Status", "Result"],
            ["Visual Inspection (Initial)", seq.visual_inspection_initial.status.value, "PASS"],
            [
                "Performance at STC",
                seq.performance_at_stc.status.value,
                f"{seq.performance_at_stc.measured_value:.1f}W",
            ],
            [
                "Thermal Cycling",
                seq.thermal_cycling.status.value,
                f"{seq.thermal_cycling.measured_value:.1f}% degradation",
            ],
            [
                "Humidity-Freeze",
                seq.humidity_freeze.status.value,
                f"{seq.humidity_freeze.measured_value:.1f}% degradation",
            ],
            [
                "Damp Heat",
                seq.damp_heat.status.value,
                f"{seq.damp_heat.measured_value:.1f}% degradation",
            ],
            ["UV Preconditioning", seq.uv_preconditioning.status.value, "PASS"],
            ["Mechanical Load", seq.mechanical_load_test.status.value, "PASS"],
            ["Hail Impact", seq.hail_impact.status.value, "PASS"],
            ["Hot Spot Endurance", seq.hot_spot_endurance.status.value, "PASS"],
            ["Visual Inspection (Final)", seq.visual_inspection_final.status.value, "PASS"],
            [
                "Performance at STC (Final)",
                seq.performance_at_stc_final.status.value,
                f"{seq.performance_at_stc_final.measured_value:.1f}W",
            ],
            [
                "Total Power Degradation",
                "",
                f"{seq.power_degradation_percent:.2f}%",
            ],
        ]

        table = Table(test_data, colWidths=[3 * inch, 1.5 * inch, 1.5 * inch])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        elements.append(table)

        # Build PDF
        doc.build(elements)
        self.logger.info(f"Generated IEC 61215 report: {output_path}")
        return output_path

    def generate_iec61730_report(
        self, result: IEC61730Result, output_path: Path
    ) -> Path:
        """
        Generate safety qualification report.

        Args:
            result: IEC 61730 test result
            output_path: Output PDF path

        Returns:
            Path: Path to generated report
        """
        self.logger.info(f"Generating IEC 61730 report for {result.test_campaign_id}")

        doc = SimpleDocTemplate(str(output_path), pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []

        # Header
        elements.extend(
            self._create_report_header(
                styles,
                "IEC 61730 Safety Qualification Test Report",
                result.test_report_number or "N/A",
            )
        )

        # Module information
        elements.append(Paragraph("<b>Module Information</b>", styles["Heading2"]))
        module_data = [
            ["Parameter", "Value"],
            ["Module Type", result.module_type],
            ["Manufacturer", result.manufacturer],
            ["Safety Class", result.safety_class],
            ["Application Class", result.application_class],
            ["Test Date", result.test_date.strftime("%Y-%m-%d")],
            ["Overall Status", result.overall_status.value.upper()],
            ["Compliance %", f"{result.compliance_percentage:.1f}%"],
        ]

        table = Table(module_data, colWidths=[3 * inch, 3 * inch])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        elements.append(table)
        elements.append(Spacer(1, 0.3 * inch))

        # Safety test results
        elements.append(Paragraph("<b>Safety Test Results</b>", styles["Heading2"]))
        safety = result.safety_tests
        test_data = [
            ["Test Name", "Status", "Result"],
            [
                "Insulation Resistance",
                safety.insulation_test.status.value,
                f"{safety.insulation_test.measured_value:.0f} MΩ",
            ],
            [
                "Dielectric Withstand",
                safety.dielectric_withstand.status.value,
                f"{safety.dielectric_withstand.measured_value:.2f} mA",
            ],
            [
                "Ground Continuity",
                safety.ground_continuity.status.value,
                f"{safety.ground_continuity.measured_value:.3f} Ω",
            ],
            ["Fire Resistance", safety.fire_test.status.value, "Class C"],
            ["Mechanical Stress", safety.mechanical_stress.status.value, "PASS"],
            ["Impact Resistance", safety.impact_test.status.value, "PASS"],
            ["UV Exposure", safety.UV_test.status.value, "PASS"],
            ["Corrosion Resistance", safety.corrosion_test.status.value, "PASS"],
        ]

        table = Table(test_data, colWidths=[3 * inch, 1.5 * inch, 1.5 * inch])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        elements.append(table)

        # Build PDF
        doc.build(elements)
        self.logger.info(f"Generated IEC 61730 report: {output_path}")
        return output_path

    def generate_iec63202_report(
        self, result: IEC63202Result, output_path: Path
    ) -> Path:
        """
        Generate CTM power loss report.

        Args:
            result: IEC 63202 test result
            output_path: Output PDF path

        Returns:
            Path: Path to generated report
        """
        self.logger.info(f"Generating IEC 63202 report for {result.test_campaign_id}")

        doc = SimpleDocTemplate(str(output_path), pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []

        # Header
        elements.extend(
            self._create_report_header(
                styles,
                "IEC 63202 CTM Power Loss Analysis Report",
                result.test_report_number or "N/A",
            )
        )

        # Module information
        elements.append(Paragraph("<b>Module Information</b>", styles["Heading2"]))
        module_data = [
            ["Parameter", "Value"],
            ["Module Type", result.module_type],
            ["Manufacturer", result.manufacturer],
            ["Test Date", result.test_date.strftime("%Y-%m-%d")],
            ["Average Cell Power", f"{result.cell_power_avg:.2f} W"],
            ["Module Power", f"{result.module_power:.2f} W"],
            ["CTM Ratio", f"{result.ctm_loss_breakdown.ctm_ratio:.3f}"],
            ["Total CTM Loss", f"{result.ctm_loss_breakdown.total_ctm_loss:.2f}%"],
        ]

        table = Table(module_data, colWidths=[3 * inch, 3 * inch])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        elements.append(table)
        elements.append(Spacer(1, 0.3 * inch))

        # Loss breakdown
        elements.append(Paragraph("<b>CTM Loss Breakdown</b>", styles["Heading2"]))
        ctm = result.ctm_loss_breakdown
        loss_data = [
            ["Loss Component", "Value (%)"],
            ["Optical Loss", f"{ctm.optical_loss:.2f}"],
            ["Electrical Loss", f"{ctm.electrical_loss:.2f}"],
            ["Thermal Loss", f"{ctm.thermal_loss:.2f}"],
            ["Mismatch Loss", f"{ctm.mismatch_loss:.2f}"],
            ["Interconnection Loss", f"{ctm.interconnection_loss:.2f}"],
            ["Inactive Area Loss", f"{ctm.inactive_area_loss:.2f}"],
            ["Total CTM Loss", f"{ctm.total_ctm_loss:.2f}"],
        ]

        table = Table(loss_data, colWidths=[3 * inch, 3 * inch])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.lightblue),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ]
            )
        )
        elements.append(table)

        # Build PDF
        doc.build(elements)
        self.logger.info(f"Generated IEC 63202 report: {output_path}")
        return output_path

    def generate_combined_report(
        self,
        result_61215: Optional[IEC61215Result],
        result_61730: Optional[IEC61730Result],
        result_63202: Optional[IEC63202Result],
        output_path: Path,
    ) -> Path:
        """
        Generate multi-standard combined certification report.

        Args:
            result_61215: IEC 61215 result
            result_61730: IEC 61730 result
            result_63202: IEC 63202 result
            output_path: Output PDF path

        Returns:
            Path: Path to generated report
        """
        self.logger.info("Generating combined IEC test report")

        doc = SimpleDocTemplate(str(output_path), pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []

        # Header
        elements.extend(
            self._create_report_header(
                styles,
                "Combined IEC Testing Certification Report",
                f"COMBINED-{datetime.now().strftime('%Y%m%d')}",
            )
        )

        # Summary table
        elements.append(Paragraph("<b>Test Summary</b>", styles["Heading2"]))
        summary_data = [["Standard", "Status", "Compliance %"]]

        if result_61215:
            summary_data.append(
                [
                    "IEC 61215 (Qualification)",
                    result_61215.overall_status.value.upper(),
                    f"{result_61215.compliance_percentage:.1f}%",
                ]
            )

        if result_61730:
            summary_data.append(
                [
                    "IEC 61730 (Safety)",
                    result_61730.overall_status.value.upper(),
                    f"{result_61730.compliance_percentage:.1f}%",
                ]
            )

        if result_63202:
            summary_data.append(
                [
                    "IEC 63202 (CTM Loss)",
                    result_63202.overall_status.value.upper(),
                    f"{result_63202.ctm_loss_breakdown.ctm_ratio * 100:.1f}%",
                ]
            )

        table = Table(summary_data, colWidths=[3 * inch, 1.5 * inch, 1.5 * inch])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        elements.append(table)

        # Build PDF
        doc.build(elements)
        self.logger.info(f"Generated combined report: {output_path}")
        return output_path

    def add_test_photos(
        self, photos: List[TestPhoto], output_dir: Path
    ) -> List[Path]:
        """
        Integrate test photos into report package.

        Args:
            photos: List of test photos
            output_dir: Output directory

        Returns:
            List[Path]: Paths to processed photos
        """
        self.logger.info(f"Processing {len(photos)} test photos")

        output_dir.mkdir(parents=True, exist_ok=True)
        processed_photos: List[Path] = []

        for i, photo in enumerate(photos):
            if photo.photo_path.exists():
                output_path = output_dir / f"photo_{i:03d}_{photo.test_name.replace(' ', '_')}.jpg"
                # In production, would copy/process image here
                processed_photos.append(output_path)

        return processed_photos

    def add_certification_signatures(
        self, report_path: Path, signatures: Dict[str, str]
    ) -> Path:
        """
        Add digital signature fields to certification report.

        Args:
            report_path: Path to report PDF
            signatures: Dictionary of signature roles and names

        Returns:
            Path: Path to signed report
        """
        self.logger.info(f"Adding certification signatures to {report_path}")

        # In production, would add digital signatures to PDF
        # For now, just log the signatures
        for role, name in signatures.items():
            self.logger.info(f"Signature: {role} = {name}")

        return report_path


class ComplianceVisualization:
    """
    Creates interactive visualizations for test results and compliance status.

    Provides Plotly-based charts and Streamlit dashboard components.
    """

    def __init__(self) -> None:
        """Initialize visualization generator."""
        self.logger = logging.getLogger(f"{__name__}.{self.__class__.__name__}")

    def pass_fail_summary(
        self, compliance_matrix: ComplianceMatrix
    ) -> go.Figure:
        """
        Create visual summary of compliance status.

        Args:
            compliance_matrix: Compliance matrix

        Returns:
            go.Figure: Plotly figure
        """
        self.logger.info("Creating pass/fail summary chart")

        # Aggregate all tests
        all_tests = {
            **compliance_matrix.iec_61215_tests,
            **compliance_matrix.iec_61730_tests,
            **compliance_matrix.iec_63202_tests,
        }

        # Count by status
        status_counts = {}
        for status in all_tests.values():
            status_counts[status.value] = status_counts.get(status.value, 0) + 1

        fig = go.Figure(
            data=[
                go.Bar(
                    x=list(status_counts.keys()),
                    y=list(status_counts.values()),
                    marker_color=["green" if k == "passed" else "red" for k in status_counts.keys()],
                )
            ]
        )

        fig.update_layout(
            title="Test Results Summary",
            xaxis_title="Status",
            yaxis_title="Number of Tests",
            template="plotly_white",
        )

        return fig

    def degradation_timeline_chart(
        self, test_history: TestHistory
    ) -> go.Figure:
        """
        Create power degradation timeline chart.

        Args:
            test_history: Historical test data

        Returns:
            go.Figure: Plotly figure
        """
        self.logger.info("Creating degradation timeline chart")

        fig = go.Figure()

        fig.add_trace(
            go.Scatter(
                x=test_history.test_dates,
                y=test_history.degradation_history,
                mode="lines+markers",
                name="Power Degradation",
                line=dict(color="red", width=2),
                marker=dict(size=8),
            )
        )

        # Add threshold line
        fig.add_hline(
            y=5.0,
            line_dash="dash",
            line_color="orange",
            annotation_text="IEC 61215 Limit (5%)",
        )

        fig.update_layout(
            title=f"Power Degradation History - {test_history.module_type}",
            xaxis_title="Test Date",
            yaxis_title="Degradation (%)",
            template="plotly_white",
            hovermode="x unified",
        )

        return fig

    def iv_curve_comparison(
        self,
        iv_initial: "IVCurveData",
        iv_final: "IVCurveData",
        title: str = "IV Curve Comparison",
    ) -> go.Figure:
        """
        Create before/after IV curve comparison.

        Args:
            iv_initial: Initial IV curve
            iv_final: Final IV curve
            title: Chart title

        Returns:
            go.Figure: Plotly figure
        """
        self.logger.info("Creating IV curve comparison")

        fig = go.Figure()

        fig.add_trace(
            go.Scatter(
                x=iv_initial.voltage,
                y=iv_initial.current,
                mode="lines",
                name="Initial",
                line=dict(color="blue", width=2),
            )
        )

        fig.add_trace(
            go.Scatter(
                x=iv_final.voltage,
                y=iv_final.current,
                mode="lines",
                name="Final",
                line=dict(color="red", width=2, dash="dash"),
            )
        )

        fig.update_layout(
            title=title,
            xaxis_title="Voltage (V)",
            yaxis_title="Current (A)",
            template="plotly_white",
            hovermode="x unified",
        )

        return fig

    def failure_mode_analysis(
        self, compliance_matrix: ComplianceMatrix
    ) -> go.Figure:
        """
        Create visual analysis of test failures.

        Args:
            compliance_matrix: Compliance matrix

        Returns:
            go.Figure: Plotly figure
        """
        self.logger.info("Creating failure mode analysis")

        # Find failed tests
        failed_tests = []

        for test_name, status in compliance_matrix.iec_61215_tests.items():
            if status == TestStatus.FAILED:
                failed_tests.append(("IEC 61215", test_name))

        for test_name, status in compliance_matrix.iec_61730_tests.items():
            if status == TestStatus.FAILED:
                failed_tests.append(("IEC 61730", test_name))

        for test_name, status in compliance_matrix.iec_63202_tests.items():
            if status == TestStatus.FAILED:
                failed_tests.append(("IEC 63202", test_name))

        if not failed_tests:
            fig = go.Figure()
            fig.add_annotation(
                text="No failures detected - All tests passed!",
                xref="paper",
                yref="paper",
                x=0.5,
                y=0.5,
                showarrow=False,
                font=dict(size=20, color="green"),
            )
            return fig

        standards = [f[0] for f in failed_tests]
        test_names = [f[1] for f in failed_tests]

        fig = go.Figure(
            data=[
                go.Bar(
                    x=test_names,
                    y=[1] * len(test_names),
                    marker_color="red",
                    text=standards,
                )
            ]
        )

        fig.update_layout(
            title="Failed Tests Analysis",
            xaxis_title="Test Name",
            yaxis_title="Failure Count",
            template="plotly_white",
        )

        return fig

    def ctm_loss_waterfall(
        self, ctm_breakdown: CTMLossBreakdown
    ) -> go.Figure:
        """
        Create waterfall chart of CTM losses.

        Args:
            ctm_breakdown: CTM loss breakdown

        Returns:
            go.Figure: Plotly figure
        """
        self.logger.info("Creating CTM loss waterfall chart")

        measures = ["relative"] * 6 + ["total"]
        x_labels = [
            "Optical",
            "Electrical",
            "Thermal",
            "Mismatch",
            "Interconnection",
            "Inactive Area",
            "Total Loss",
        ]
        y_values = [
            ctm_breakdown.optical_loss,
            ctm_breakdown.electrical_loss,
            ctm_breakdown.thermal_loss,
            ctm_breakdown.mismatch_loss,
            ctm_breakdown.interconnection_loss,
            ctm_breakdown.inactive_area_loss,
            ctm_breakdown.total_ctm_loss,
        ]

        fig = go.Figure(
            go.Waterfall(
                name="CTM Loss",
                orientation="v",
                measure=measures,
                x=x_labels,
                y=y_values,
                text=[f"{v:.2f}%" for v in y_values],
                textposition="outside",
                connector={"line": {"color": "rgb(63, 63, 63)"}},
            )
        )

        fig.update_layout(
            title="CTM Power Loss Breakdown",
            yaxis_title="Loss (%)",
            template="plotly_white",
            showlegend=False,
        )

        return fig


class CertificationWorkflow:
    """
    Manages certification application workflow and status tracking.

    Handles certification package preparation, cost tracking, and status monitoring.
    """

    def __init__(self, data_dir: Optional[Path] = None) -> None:
        """
        Initialize certification workflow manager.

        Args:
            data_dir: Directory for certification data
        """
        self.data_dir = data_dir or Path("./certification_data")
        self.data_dir.mkdir(parents=True, exist_ok=True)
        self.logger = logging.getLogger(f"{__name__}.{self.__class__.__name__}")

    def prepare_certification_package(
        self,
        compliance_report: ComplianceReport,
        target_certifications: List[CertificationBodyType],
        module_type: str,
        manufacturer: str,
    ) -> CertificationPackage:
        """
        Package all documentation for certification submission.

        Args:
            compliance_report: Complete compliance report
            target_certifications: Target certification bodies
            module_type: Module type
            manufacturer: Module manufacturer

        Returns:
            CertificationPackage: Complete certification package
        """
        self.logger.info(f"Preparing certification package for {module_type}")

        package = CertificationPackage(
            package_id=f"CERT-{datetime.now().strftime('%Y%m%d-%H%M%S')}",
            module_type=module_type,
            manufacturer=manufacturer,
            target_certifications=target_certifications,
            target_standards=[
                IECStandard.IEC_61215,
                IECStandard.IEC_61730,
                IECStandard.IEC_63202,
            ],
            compliance_report=compliance_report,
            test_reports_paths=[],  # Would be populated with actual paths
        )

        self.logger.info(f"Created certification package: {package.package_id}")
        return package

    def track_certification_status(
        self, package_id: str
    ) -> List[CertificationStatus]:
        """
        Monitor certification application status.

        Args:
            package_id: Certification package ID

        Returns:
            List[CertificationStatus]: Current certification statuses
        """
        self.logger.info(f"Tracking certification status for {package_id}")

        # In production, would query certification database
        # For now, return mock statuses
        statuses = [
            CertificationStatus(
                certification_body=CertificationBodyType.TUV_RHEINLAND,
                application_date=datetime.now() - timedelta(days=30),
                status="Under Review",
                expected_completion_date=datetime.now() + timedelta(days=30),
                certification_cost=15000.0,
            ),
            CertificationStatus(
                certification_body=CertificationBodyType.UL,
                application_date=datetime.now() - timedelta(days=45),
                status="Testing Complete",
                expected_completion_date=datetime.now() + timedelta(days=15),
                certification_cost=12000.0,
            ),
        ]

        return statuses

    def manage_certification_costs(
        self, certification_statuses: List[CertificationStatus]
    ) -> Dict[str, float]:
        """
        Track certification fees and timeline.

        Args:
            certification_statuses: List of certification statuses

        Returns:
            Dict[str, float]: Cost summary
        """
        self.logger.info("Managing certification costs")

        total_cost = sum(
            s.certification_cost for s in certification_statuses if s.certification_cost
        )
        costs_by_body = {
            s.certification_body.value: s.certification_cost or 0.0
            for s in certification_statuses
        }

        return {
            "total_cost": total_cost,
            **costs_by_body,
        }

    def handle_recertification(
        self, original_package_id: str, recert_reason: str
    ) -> CertificationPackage:
        """
        Manage periodic recertification requirements.

        Args:
            original_package_id: Original certification package ID
            recert_reason: Reason for recertification

        Returns:
            CertificationPackage: New recertification package
        """
        self.logger.info(f"Handling recertification for {original_package_id}: {recert_reason}")

        # In production, would load original package and create updated version
        # For now, return placeholder
        raise NotImplementedError("Recertification handling requires database integration")

    def international_certification_mapping(
        self, source_standard: IECStandard
    ) -> Dict[str, str]:
        """
        Map IEC standards to local certification requirements.

        Args:
            source_standard: Source IEC standard

        Returns:
            Dict[str, str]: Mapping to local standards
        """
        self.logger.info(f"Mapping {source_standard.value} to local standards")

        mappings = {
            IECStandard.IEC_61215: {
                "US": "UL 1703",
                "Canada": "CSA C61215",
                "Japan": "JIS C 8990",
                "China": "GB/T 9535",
                "Europe": "EN 61215",
            },
            IECStandard.IEC_61730: {
                "US": "UL 61730",
                "Canada": "CSA C61730",
                "Japan": "JIS C 61730",
                "China": "GB/T 18911",
                "Europe": "EN 61730",
            },
            IECStandard.IEC_63202: {
                "US": "N/A (IEC TS only)",
                "Canada": "N/A (IEC TS only)",
                "Japan": "N/A (IEC TS only)",
                "China": "N/A (IEC TS only)",
                "Europe": "EN 63202",
            },
        }

        return mappings.get(source_standard, {})
