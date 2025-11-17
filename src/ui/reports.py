"""Comprehensive Reports Module.

This module provides comprehensive reporting capabilities including:
- PDF report generation
- Excel export
- Sensitivity analysis tables
- P50/P90/P99 probabilistic analysis
"""

from datetime import datetime
from io import BytesIO
from typing import Dict, List, Optional
import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    PageBreak,
    Image,
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..models.eya_models import (
    ProjectInfo,
    SystemConfiguration,
    PerformanceMetrics,
    FinancialMetrics,
    ProbabilisticAnalysis,
    SensitivityAnalysis,
)
from ..modules.B06_energy_yield_analysis.analyzer import EnergyYieldAnalyzer


class ComprehensiveReports:
    """Comprehensive reporting engine for Energy Yield Analysis.

    This class provides advanced reporting capabilities including:
    - Professional PDF report generation
    - Excel export with multiple sheets
    - Sensitivity analysis tables
    - P50/P90/P99 probabilistic analysis

    Attributes:
        project_info: Project information
        system_config: System configuration
        analyzer: Energy yield analyzer instance
    """

    def __init__(
        self,
        project_info: ProjectInfo,
        system_config: SystemConfiguration,
        analyzer: EnergyYieldAnalyzer,
    ):
        """Initialize the comprehensive reports generator.

        Args:
            project_info: Project information
            system_config: System configuration
            analyzer: Energy yield analyzer instance
        """
        self.project_info = project_info
        self.system_config = system_config
        self.analyzer = analyzer

    def eya_pdf_generator(
        self,
        annual_energy: float,
        performance_metrics: PerformanceMetrics,
        financial_metrics: Optional[FinancialMetrics] = None,
    ) -> BytesIO:
        """Generate comprehensive EYA PDF report.

        Args:
            annual_energy: Annual energy production in kWh
            performance_metrics: Performance metrics
            financial_metrics: Optional financial metrics

        Returns:
            BytesIO buffer containing PDF report
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.75 * inch)

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=24,
            textColor=colors.HexColor("#1f77b4"),
            spaceAfter=30,
            alignment=TA_CENTER,
        )
        heading_style = ParagraphStyle(
            "CustomHeading",
            parent=styles["Heading2"],
            fontSize=14,
            textColor=colors.HexColor("#2c3e50"),
            spaceAfter=12,
            spaceBefore=12,
        )

        # Build document
        story = []

        # Title Page
        story.append(Paragraph("Energy Yield Analysis Report", title_style))
        story.append(Spacer(1, 0.3 * inch))

        # Project Information
        story.append(Paragraph("Project Information", heading_style))
        project_data = [
            ["Project Name:", self.project_info.project_name],
            ["Location:", self.project_info.location],
            ["Coordinates:", f"{self.project_info.latitude:.4f}°, {self.project_info.longitude:.4f}°"],
            ["Commissioning Date:", self.project_info.commissioning_date.strftime("%Y-%m-%d")],
            ["Project Lifetime:", f"{self.project_info.project_lifetime} years"],
        ]
        project_table = Table(project_data, colWidths=[2.5 * inch, 4 * inch])
        project_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#ecf0f1")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ])
        )
        story.append(project_table)
        story.append(Spacer(1, 0.3 * inch))

        # System Configuration
        story.append(Paragraph("System Configuration", heading_style))
        system_data = [
            ["DC Capacity:", f"{self.system_config.capacity_dc:.2f} kWp"],
            ["AC Capacity:", f"{self.system_config.capacity_ac:.2f} kWac"],
            ["Module Type:", self.system_config.module_type.value],
            ["Module Count:", f"{self.system_config.module_count:,}"],
            ["Inverter Efficiency:", f"{self.system_config.inverter_efficiency * 100:.2f}%"],
            ["Tilt/Azimuth:", f"{self.system_config.tilt_angle:.1f}° / {self.system_config.azimuth_angle:.1f}°"],
        ]
        system_table = Table(system_data, colWidths=[2.5 * inch, 4 * inch])
        system_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#ecf0f1")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ])
        )
        story.append(system_table)
        story.append(Spacer(1, 0.3 * inch))

        # Energy Production
        story.append(Paragraph("Annual Energy Production", heading_style))
        specific_yield = annual_energy / self.system_config.capacity_dc
        capacity_factor = annual_energy / (self.system_config.capacity_ac * 8760)

        energy_data = [
            ["Annual AC Energy:", f"{annual_energy:,.0f} kWh"],
            ["Specific Yield:", f"{specific_yield:.2f} kWh/kWp"],
            ["Capacity Factor:", f"{capacity_factor * 100:.2f}%"],
        ]
        energy_table = Table(energy_data, colWidths=[2.5 * inch, 4 * inch])
        energy_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#ecf0f1")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ])
        )
        story.append(energy_table)
        story.append(Spacer(1, 0.3 * inch))

        # Performance Metrics
        story.append(Paragraph("Performance Analysis", heading_style))
        perf_data = [
            ["Performance Ratio:", f"{performance_metrics.performance_ratio * 100:.2f}%"],
            ["Reference Yield:", f"{performance_metrics.reference_yield:.2f} kWh/kWp"],
            ["Array Yield:", f"{performance_metrics.array_yield:.2f} kWh/kWp"],
            ["Final Yield:", f"{performance_metrics.final_yield:.2f} kWh/kWp"],
            ["System Losses:", f"{performance_metrics.system_losses:.2f} kWh/kWp"],
        ]
        perf_table = Table(perf_data, colWidths=[2.5 * inch, 4 * inch])
        perf_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#ecf0f1")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ])
        )
        story.append(perf_table)
        story.append(Spacer(1, 0.3 * inch))

        # Loss Analysis
        story.append(Paragraph("Loss Analysis", heading_style))
        losses = self.analyzer.analyze_losses()
        loss_data = [
            ["Loss Category", "Value (%)"],
            ["Soiling Loss", f"{losses['soiling_loss']:.2f}%"],
            ["Shading Loss", f"{losses['shading_loss']:.2f}%"],
            ["Temperature Loss", f"{losses['temperature_loss']:.2f}%"],
            ["Inverter Loss", f"{losses['inverter_loss']:.2f}%"],
            ["Total System Loss", f"{losses['total_loss']:.2f}%"],
        ]
        loss_table = Table(loss_data, colWidths=[3 * inch, 3.5 * inch])
        loss_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3498db")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ])
        )
        story.append(loss_table)

        # Financial Analysis (if provided)
        if financial_metrics and financial_metrics.lcoe:
            story.append(PageBreak())
            story.append(Paragraph("Financial Analysis", heading_style))

            financial_data = [
                ["CAPEX:", f"${financial_metrics.capex:,.0f}"],
                ["Annual OPEX:", f"${financial_metrics.opex_annual:,.0f}"],
                ["LCOE:", f"${financial_metrics.lcoe:.4f}/kWh"],
                ["NPV:", f"${financial_metrics.npv:,.0f}"],
                ["IRR:", f"{financial_metrics.irr * 100:.2f}%"],
                ["Payback Period:", f"{financial_metrics.payback_period:.1f} years"],
            ]
            financial_table = Table(financial_data, colWidths=[2.5 * inch, 4 * inch])
            financial_table.setStyle(
                TableStyle([
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#ecf0f1")),
                    ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ])
            )
            story.append(financial_table)

        # Footer
        story.append(Spacer(1, 0.5 * inch))
        footer_text = f"Report generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        story.append(Paragraph(footer_text, styles["Normal"]))

        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer

    def excel_export(
        self,
        annual_energy: float,
        performance_metrics: PerformanceMetrics,
        monthly_data: pd.DataFrame,
        financial_metrics: Optional[FinancialMetrics] = None,
    ) -> BytesIO:
        """Export comprehensive data to Excel workbook.

        Args:
            annual_energy: Annual energy production
            performance_metrics: Performance metrics
            monthly_data: Monthly performance data
            financial_metrics: Optional financial metrics

        Returns:
            BytesIO buffer containing Excel workbook
        """
        buffer = BytesIO()
        workbook = openpyxl.Workbook()

        # Remove default sheet
        workbook.remove(workbook.active)

        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Sheet 1: Project Overview
        ws_overview = workbook.create_sheet("Project Overview")
        ws_overview["A1"] = "Energy Yield Analysis Report"
        ws_overview["A1"].font = title_font
        ws_overview.merge_cells("A1:B1")

        overview_data = [
            ["", ""],
            ["Project Information", ""],
            ["Project Name", self.project_info.project_name],
            ["Location", self.project_info.location],
            ["Latitude", self.project_info.latitude],
            ["Longitude", self.project_info.longitude],
            ["Commissioning Date", self.project_info.commissioning_date.strftime("%Y-%m-%d")],
            ["", ""],
            ["System Configuration", ""],
            ["DC Capacity (kWp)", self.system_config.capacity_dc],
            ["AC Capacity (kWac)", self.system_config.capacity_ac],
            ["Module Type", self.system_config.module_type.value],
            ["Module Count", self.system_config.module_count],
            ["Inverter Efficiency (%)", self.system_config.inverter_efficiency * 100],
            ["Tilt Angle (°)", self.system_config.tilt_angle],
            ["Azimuth Angle (°)", self.system_config.azimuth_angle],
        ]

        for row_idx, (label, value) in enumerate(overview_data, start=1):
            ws_overview.cell(row=row_idx, column=1, value=label)
            ws_overview.cell(row=row_idx, column=2, value=value)

        ws_overview.column_dimensions["A"].width = 30
        ws_overview.column_dimensions["B"].width = 25

        # Sheet 2: Energy Production
        ws_energy = workbook.create_sheet("Energy Production")
        ws_energy["A1"] = "Annual Energy Production"
        ws_energy["A1"].font = title_font

        specific_yield = annual_energy / self.system_config.capacity_dc
        capacity_factor = annual_energy / (self.system_config.capacity_ac * 8760)

        energy_data = [
            ["Metric", "Value"],
            ["Annual AC Energy (kWh)", annual_energy],
            ["Specific Yield (kWh/kWp)", specific_yield],
            ["Capacity Factor (%)", capacity_factor * 100],
            ["Performance Ratio (%)", performance_metrics.performance_ratio * 100],
        ]

        for row_idx, row_data in enumerate(energy_data, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_energy.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 3:  # Header row
                    cell.fill = header_fill
                    cell.font = header_font
                cell.border = border

        ws_energy.column_dimensions["A"].width = 30
        ws_energy.column_dimensions["B"].width = 20

        # Sheet 3: Monthly Production
        ws_monthly = workbook.create_sheet("Monthly Production")
        ws_monthly["A1"] = "Monthly Energy Production"
        ws_monthly["A1"].font = title_font

        # Write headers
        headers = ["Month", "DC Energy (kWh)", "AC Energy (kWh)", "Specific Yield (kWh/kWp)", "Capacity Factor (%)"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_monthly.cell(row=3, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border

        # Write data
        for row_idx, row in enumerate(monthly_data.itertuples(index=False), start=4):
            ws_monthly.cell(row=row_idx, column=1, value=row.month).border = border
            ws_monthly.cell(row=row_idx, column=2, value=row.dc_energy).border = border
            ws_monthly.cell(row=row_idx, column=3, value=row.ac_energy).border = border
            ws_monthly.cell(row=row_idx, column=4, value=row.specific_yield).border = border
            ws_monthly.cell(row=row_idx, column=5, value=row.capacity_factor * 100).border = border

        for col in ["A", "B", "C", "D", "E"]:
            ws_monthly.column_dimensions[col].width = 20

        # Sheet 4: Loss Analysis
        ws_losses = workbook.create_sheet("Loss Analysis")
        ws_losses["A1"] = "System Loss Analysis"
        ws_losses["A1"].font = title_font

        losses = self.analyzer.analyze_losses()
        loss_headers = ["Loss Type", "Loss (%)"]

        for col_idx, header in enumerate(loss_headers, start=1):
            cell = ws_losses.cell(row=3, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border

        loss_items = [
            ("Soiling Loss", losses["soiling_loss"]),
            ("Shading Loss", losses["shading_loss"]),
            ("Snow Loss", losses["snow_loss"]),
            ("Mismatch Loss", losses["mismatch_loss"]),
            ("Wiring Loss", losses["wiring_loss"]),
            ("Temperature Loss", losses["temperature_loss"]),
            ("Inverter Loss", losses["inverter_loss"]),
            ("Total System Loss", losses["total_loss"]),
        ]

        for row_idx, (loss_name, loss_value) in enumerate(loss_items, start=4):
            ws_losses.cell(row=row_idx, column=1, value=loss_name).border = border
            ws_losses.cell(row=row_idx, column=2, value=loss_value).border = border

        ws_losses.column_dimensions["A"].width = 25
        ws_losses.column_dimensions["B"].width = 15

        # Sheet 5: Financial Analysis (if provided)
        if financial_metrics and financial_metrics.lcoe:
            ws_financial = workbook.create_sheet("Financial Analysis")
            ws_financial["A1"] = "Financial Metrics"
            ws_financial["A1"].font = title_font

            financial_data = [
                ["Metric", "Value"],
                ["CAPEX ($)", financial_metrics.capex],
                ["Annual OPEX ($)", financial_metrics.opex_annual],
                ["Energy Price ($/kWh)", financial_metrics.energy_price],
                ["LCOE ($/kWh)", financial_metrics.lcoe],
                ["NPV ($)", financial_metrics.npv],
                ["IRR (%)", financial_metrics.irr * 100 if financial_metrics.irr else 0],
                ["Payback Period (years)", financial_metrics.payback_period],
            ]

            for row_idx, row_data in enumerate(financial_data, start=3):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws_financial.cell(row=row_idx, column=col_idx, value=value)
                    if row_idx == 3:
                        cell.fill = header_fill
                        cell.font = header_font
                    cell.border = border

            ws_financial.column_dimensions["A"].width = 30
            ws_financial.column_dimensions["B"].width = 20

        # Save workbook
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    def sensitivity_analysis_tables(
        self, base_annual_energy: float, parameters: List[tuple]
    ) -> pd.DataFrame:
        """Generate sensitivity analysis tables for multiple parameters.

        Args:
            base_annual_energy: Base case annual energy
            parameters: List of tuples (parameter_name, base_value, variation_pct)

        Returns:
            DataFrame with sensitivity analysis results
        """
        results = []

        for param_name, base_value, variation_pct in parameters:
            sensitivity = self.analyzer.perform_sensitivity_analysis(
                base_annual_energy, param_name, base_value, variation_pct
            )

            for param_value, energy_value in sensitivity.results.items():
                pct_change = ((param_value - base_value) / base_value) * 100
                energy_change = ((energy_value - base_annual_energy) / base_annual_energy) * 100

                results.append({
                    "Parameter": param_name,
                    "Base Value": base_value,
                    "Test Value": param_value,
                    "Change (%)": pct_change,
                    "Annual Energy (kWh)": energy_value,
                    "Energy Change (%)": energy_change,
                })

        return pd.DataFrame(results)

    def p50_p90_p99_analysis(
        self, base_annual_energy: float, uncertainty_pct: float = 10.0
    ) -> Dict:
        """Generate P50/P90/P99 probabilistic analysis.

        Args:
            base_annual_energy: Base case annual energy
            uncertainty_pct: Total uncertainty percentage

        Returns:
            Dictionary with probabilistic analysis results
        """
        # Calculate probabilistic analysis
        prob_analysis = self.analyzer.calculate_probabilistic_yield(
            base_annual_energy, uncertainty_pct
        )

        # Create analysis table
        analysis_data = {
            "Exceedance Probability": {
                "P99 (99%)": prob_analysis.p99,
                "P90 (90%)": prob_analysis.p90,
                "P75 (75%)": prob_analysis.p75,
                "P50 (50% - Median)": prob_analysis.p50,
                "Mean": prob_analysis.mean,
            },
            "Statistics": {
                "Standard Deviation (kWh)": prob_analysis.std_dev,
                "Coefficient of Variation (%)": (prob_analysis.std_dev / prob_analysis.mean) * 100,
            },
            "Confidence Intervals": prob_analysis.confidence_intervals,
            "Analysis": {
                "P90/P50 Ratio": prob_analysis.p90 / prob_analysis.p50 if prob_analysis.p50 > 0 else 0,
                "P99/P50 Ratio": prob_analysis.p99 / prob_analysis.p50 if prob_analysis.p50 > 0 else 0,
                "Upside (P90 - P50)": prob_analysis.p90 - prob_analysis.p50,
                "Downside (P50 - P99)": prob_analysis.p50 - prob_analysis.p99,
            },
        }

        # Create DataFrame for easy viewing
        df_data = []
        for probability, value in analysis_data["Exceedance Probability"].items():
            df_data.append({
                "Probability Level": probability,
                "Annual Energy (kWh)": value,
                "% of P50": (value / prob_analysis.p50) * 100 if prob_analysis.p50 > 0 else 0,
            })

        analysis_data["table"] = pd.DataFrame(df_data)

        return analysis_data
